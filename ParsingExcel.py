# -*- coding: utf-8 -*-
"""
Created on Fri Apr 13 14:55:47 2018
Metadata to GIS files
@author: daniel.scott
"""
from openpyxl import load_workbook

wb = load_workbook(r'W:/daniel.scott/Search tool/updatedV04DatasourcesV2.xlsx',read_only = False)
sheet = wb["sheet6"]
geology = ["Geology","litho","geol","Lithology","Seccion","Alteration","alt","Cover","Struc","falla","fault","Intrusives","sections","Observation","Geochronology","geochron","Mineralization","Thin"]
geochemistry = ["ASD","Sampling","Assaying","XRF","Greenrocks","Fertility","geochem","MEMS"]
geog = ["Geography","road","river","cities","pueblos"]
geophysics = ["Geophys","Mag","IP","Seis","Grav","radiome"]

#working script need to add 
#starting from row 2 - excluding headers until the maximum row with data
for rowNum in range(2, sheet.max_row):
    #Look in column B (column=2) for item in geology if found write in column 3 (C) the tag
    if any(x in sheet.cell(row=rowNum, column=2).value for x in geology):
        print sheet.cell(row=rowNum, column=2).value
        sheet.cell(row = rowNum,column=3).value = "Geology"
    elif any(x in sheet.cell(row=rowNum, column=2).value for x in geochemistry):
        print sheet.cell(row=rowNum, column=2).value
        sheet.cell(row = rowNum,column=3).value = "Geochem"
    elif any(x in sheet.cell(row=rowNum, column=2).value for x in geog):
        print sheet.cell(row=rowNum, column=2).value
        sheet.cell(row = rowNum,column=3).value = "Geog"
    elif any(x in sheet.cell(row=rowNum, column=2).value for x in geophysics):
        print sheet.cell(row=rowNum, column=2).value
        sheet.cell(row = rowNum,column=3).value = "Geophysics"
    elif any(x in sheet.cell(row=rowNum, column=1).value for x in geology):
        print sheet.cell(row=rowNum, column=2).value
        sheet.cell(row = rowNum,column=4).value = "Geology"
    elif any(x in sheet.cell(row=rowNum, column=1).value for x in geochemistry):
        print sheet.cell(row=rowNum, column=2).value
        sheet.cell(row = rowNum,column=4).value = "Geochem"
    elif any(x in sheet.cell(row=rowNum, column=1).value for x in geog):
        print sheet.cell(row=rowNum, column=2).value
        sheet.cell(row = rowNum,column=4).value = "Geog"
    elif any(x in sheet.cell(row=rowNum, column=1).value for x in geophysics):
        print sheet.cell(row=rowNum, column=2).value
        sheet.cell(row = rowNum,column=4).value = "Geophysics"
#has to create new copy of worksheet
wb.save(r'W:/daniel.scott/Search tool/updatedV04Datasources_ScriptRun99.xlsx')