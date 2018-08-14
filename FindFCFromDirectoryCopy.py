# -*- coding: utf-8 -*-
"""
Created on Wed Apr 18 16:27:03 2018
This script is run in arccatalog to take files that have been previously mapped with the Directory column, and moves them to a FGDB renaming them with the filename column, then creates tags for these files according to dictionaries created and adds them to the FC now in the FGDB
@author: daniel.scott
"""
import os, arcpy, unidecode, sys
from openpyxl import load_workbook
import arcpy_metadata as md

gdb_name = "GDBProject.gdb"
#workspace something like r"Q:\Search tool"
ws = r"Q:\<INSERT DIRECTORY>"
gdbpath = os.path.join(ws, gdb_name)

#Input the excel which contains the directory of the files you want to copy i.e. r'C:\Datasources.xlsx'
wb = load_workbook(r'C:\datasources.xlsx',read_only = False)

#specify the sheet name, below Firstrun is the sheet
sheet = wb["Firstrun"]

#for row take the metadata
#
for rowNum in range(5, sheet.max_row):
    datasource = sheet.cell(row=rowNum, column=1).value #directory to file
    print datasource
    if arcpy.Exists(datasource):
        print datasource + " is a file"
        name = sheet.cell(row=rowNum, column=2).value # value of filename
        name = unidecode.unidecode(name)
        name = name.replace(" ", "_")
        if name.endswith(".shp"):
            name = name.split(".")[0]
        else:
            continue
        print "filename to be called" + name
        try:
            arcpy.FeatureClassToFeatureClass_conversion(datasource,gdbpath,name)
            print datasource + " has been successfulled moved to" + gdbpath + name        
        except Exception:
            e = sys.exc_info()[1]
            print(e.args[0])
            pass
    else:
        print "file does not exists or cannot access: "+ name
        pass

###for gdb files
arcpy.env.workspace = ##path to gdb that holds the files
for dataset in arcpy.ListDatasets():      
    for fc in arcpy.ListFeatureClasses("*", "ALL", dataset):
        gdb_tomove = r'C:\GDB.gdb'
        arcpy.FeatureClassToGeodatabase_conversion(fc, gdb_tomove)
