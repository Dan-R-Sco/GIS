# -*- coding: utf-8 -*-
"""
Created on Fri Apr 20 11:03:06 2018
Script that will read a excel to a defined sheet and write tags (key words) next to the directory column
@author: daniel.scott
"""

from openpyxl import load_workbook
##input excel document
wb = load_workbook(<INPUT EXCEL>,read_only = False)

#set key words that (list headers) and the words that could mean them (lists contents)  
geology = ["cover","lithology", "rock","age","litho", "geol", "geology","drill","thin","fault", "strk","min", "vein","struct","hsdec","handsample","lito","geochron","alt","halo","section", "seccion","lith"]
geochemistry = ["geochem", "mems","geochemistry", "geoch","gch", "asd","alter","ger","sampling","assaying","xrf","greenrock","fertility"]
geophysics = ["geophy", "mag", "geop","ip","seis","grav","radiome"]
geography = ["topo","road","citi","pueblo","rivers","drenaje","glaciares"]
commercial = ["access", "Concesione", "exploit","propiet","propert","communit","environmen","rights"]
subsurface = ["drill","subsurf","subsue", "sond", "dh"]
sheet = wb["Firstrun"]


for rowNum in range(5, sheet.max_row):
    cell = sheet.cell(row=rowNum, column=1).value
    try:
        lowercell = cell.lower()
    except:
        lowercell = cell
    try:
        cell = sheet.cell(row=rowNum, column=1).value 
        lowercell = cell.lower()
        if any(x in lowercell for x in geology):
            print sheet.cell(row=rowNum, column=1).value
            sheet.cell(row = rowNum,column=5).value = "Geology"
    except:
        print "unable to process" + str(rowNum)
    try:
        if any(x in lowercell for x in geochemistry):
            sheet.cell(row = rowNum,column=6).value = "Geochemistry"
    except:
        print "unable to process" + str(rowNum)
    try:
        if any(x in lowercell for x in geophysics):
            sheet.cell(row = rowNum,column=7).value = "Geophysics"
    except:
        print "unable to process" + str(rowNum)
    try:
        if any(x in lowercell for x in geography):
            sheet.cell(row = rowNum,column=8).value = "Geography"
    except:
        print "unable to process" + str(rowNum)
    try:
        if any(x in lowercell for x in commercial):
            sheet.cell(row = rowNum,column=9).value = "Commercial"
    except:
        print "unable to process" + str(rowNum)
    try:
        if any(x in lowercell for x in subsurface):
            sheet.cell(row = rowNum,column=10).value = "Subsurface"
    except:
        print "unable to process" + str(rowNum)                                                                                                                                                              

        #define where to save the xls thats been edited
wb.save(<INPUT EXCEL DIRECTORY AND FILE NAME WITH EXTENSION>)
