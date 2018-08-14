# -*- coding: utf-8 -*-
"""
Created on Fri Apr 13 14:55:47 2018
Reads an excel file and checks the directory to see if inside it has any keywords that we can associate to metadata categories and writes this to the next column
Then it tries to assign this keyword to the file
@author: daniel.scott
"""

from openpyxl import load_workbook

#Add workbook i.e. r'C:\Docs\tool\Datasources_Script.xlsx'
wb = load_workbook(<WORKSBOOK DIRECTORY AND NAME>,read_only = False)
geology = ["cover","lithology", "rock","drill","Age","litho", "geol", "geology","drill","thin","fault", "strk","Min", "vein","struct","HSDec","handsample","lito","geochron","alt","halo","section", "seccion"]
geochemistry = ["geochem", "MEM","geochemistry", "geoch","gch", "ASD","alter","GER"]
geophysics = ["geophy", "mag", "geop",""]
geography = ["topo","road","citi","pueblo","rivers","","drenaje","glaciares",""]
commercial = ["access", "Concesione", "exploit","propiet","propert","communit","environmen","rights"]
subsurface = ["drill","subsurf","subsue", "sond", "DH"]
sheet = wb["Firstrun"]

for rowNum in range(2, sheet.max_row):          
    directory ="source directory: " + sheet.cell(row=rowNum, column=1).value #change value to that of the directory
    if any(geology) in directory:
        sheet.cell(row = rowNum,column=5).value = "Geology"
    if any(geochemistry) in directory:
        sheet.cell(row = rowNum,column=6).value = "Geochemistry"
    if any(geophysics) in directory:
        sheet.cell(row = rowNum,column=7).value = "Geophysics"
    if any(geography) in directory:
        sheet.cell(row = rowNum,column=8).value = "Geography"
    if any(geophysics) in directory:
        sheet.cell(row = rowNum,column=9).value = "Commercial"
    if any(subsurface) in directory:
        sheet.cell(row = rowNum,column=10).value = "Subsurface"
            try:
                fileloc = gdbpath + name
                if arcpy.exists(fileloc):
                    print "found " + fileloc
                    metadata = md.MetadataEditor(name)
                    #add geology tag
    
                    tagdirvalue = sheet.cell(row = rowNum, column = 4).value
                    print "row number: " + rowNum + ", tag value: " + tagdirvalue  
                    metadata.tags.append(tagdirvalue)
                    #add geochem tag
                    tagdirvalue2 = sheet.cell(row = rowNum, column = 5).value
                    if tagdirvalue2 is None:
                        print "No tag for second column"
                        continue
                    else:
                        print "row number: " + rowNum + ", tag value: " + tagdirvalue2
                        metadata.tags.append(tagdirvalue2)
                    #add geophy tag
                    tagdirvalue3 = sheet.cell(row = rowNum, column = 6).value
                    if tagdirvalue3 is None:
                        pass
                    else:
                        print "row number: " + rowNum + ", tag value: " + tagdirvalue3  
                        metadata.tags.append(tagdirvalue3)
                    #add geography tag
                    tagdirvalue4 = sheet.cell(row = rowNum, column = 7).value
                    if tagdirvalue4 is None:
                        pass
                    else:
                        print "row number: " + rowNum + ", tag value: " + tagdirvalue4  
                        metadata.tags.append(tagdirvalue4)
                    #add commercial tag
                    tagdirvalue5 = sheet.cell(row = rowNum, column = 8).value
                    if tagdirvalue5 is None:
                        pass
                    else:
                        print "row number: " + rowNum + ", tag value: " + tagdirvalue5  
                        metadata.tags.append(tagdirvalue5)
                    #add subsurface tag
                    tagdirvalue6 = sheet.cell(row = rowNum, column = 9).value
                    if tagdirvalue6 is None:
                        pass
                    else:
                        print "row number: " + rowNum + ", tag value: " + tagdirvalue6
                        metadata.tags.append(tagdirvalue5)
                    
                    print "added metadata from directory for: " +tagdirvalue + tagdirvalue2 + tagdirvalue3 + tagdirvalue4 +tagdirvalue5 + tagdirvalue6 + " added to " + fileloc     
                else:
                    print "unable to find " + name   
            except:
                print "unable to add tag to " + fileloc

print datasource + " copied to EGDB as " + name 
arcpy_metadata.metadata.finish()  # save() and cleanup() as one call
#save the xls after writing to it
wb.save(<DIRECTORY AND XLS NAME>)

"""
some commands that can be used in md editor
metadata = md.MetadataEditor(path_to_fc)

title = metadata.title
metadata.title = "The new title"
#Get list items (returns list)
tags = metadata.tags
for tag in tags:
    print tag
#Change list items
metadata.tags = ["V04", "tag2"]
metadata.tags[1] = "another tag"
metadata.tags.append("new tag")
metadata.tags.remove("V04")
metadata.tags.pop()
#Get numeric items (return int or float)

min_scale = metadata.min_scale
max_scale = metadata.max_scale
#Change numeric items

metadata.min_scale = 500000
metadata.max_scale = 500

#Saving the changes back to the file
metadata.save() # save the metadata back to file.
metadata.cleanup() # remove all temporary files.
metadata.finish()  # save() and cleanup() as one call
"""

"""
#olderscript
from openpyxl import load_workbook
import arcpy_metadata as md
import glob, os

#read the xls workbook
wb = load_workbook(<XLS FILE AND DIRECTORY>,read_only = False)
#define the sheet to read
sheet = wb["alltags"]

notags = []
#cycle through each row
for rowNum in range(2, sheet.max_row):
    if sheet.cell(row = rowNum, column = 3).value is None:
        notags.append(sheet.cell(row = rowNum, column = 2).value)
        pass
    else:
        print sheet.cell(row = rowNum, column = 3).value
        fileloc = glob.glob(sheet.cell(row = rowNum, column = 2).value)
#locate file
        try:
            metadata = md.MetadataEditor(fileloc)
            #add project tag
            metadata.tags = ["V04"]
            tagdirvalue = sheet.cell(row = rowNum, column = 3).value
            print "row number: " + rowNum + ", tag value: " + tagdirvalue  
            #add directory tag
            metadata.tags.append(tagdirvalue)
            print "added metadata from directory for: " + "'" + "'" + tagdirvalue + " added to " + fileloc     
        except:
            print "unable to add tag to " + fileloc
        try:
            tagmxdvalue = sheet.cell(row = rowNum, column = 4).value
            print "row number: " + rowNum + ", tag value: " + tagmxdvalue  
            #add tag of dicipline if available
            if tagmxdvalue is not None:
                metadata.tags.append(tagmxdvalue)
                print "added metadata from mxd grouping " + "'" + "'" + tagmxdvalue + " added to " + fileloc
            else:
                pass
        except:
            print "unable to add tag to " + fileloc
metadata.finish()  # save() and cleanup() as one call
